"""
Enhanced Excel extractor with LLM-based intelligent sheet detection
"""

import os
import openpyxl
from typing import Dict, List, Any, Optional
from datetime import datetime

from models import (
    ExtractionResult, ExtractedQuestion, HierarchyStats, QuestionType,
    SheetType, GlobalContext
)
from claude_client import ClaudeStructuredClient
from parsers import HierarchicalQuestionParser, MultiLineQuestionParser
from smart_sheet_analyzer import SmartSheetAnalyzer


class EnhancedExcelExtractor:
    """Enhanced extractor with LLM-based intelligent detection"""

    def __init__(self, file_path: str, use_llm_hierarchy=True, content_sheet_name=None):
        self.use_llm_hierarchy = use_llm_hierarchy
        self.file_path = file_path
        self.workbook = None
        self.claude_client = ClaudeStructuredClient()
        self.smart_analyzer = SmartSheetAnalyzer(self.claude_client)
        self.hierarchy_parser = HierarchicalQuestionParser() if use_llm_hierarchy else None
        self.global_context = None
        self.content_sheet_name = content_sheet_name
        self.detected_content_sheet = None

    def _get_headers(self, sheet) -> List[Dict[str, str]]:
        """Get sheet headers"""
        headers = []
        for col in range(1, min(20, sheet.max_column + 1)):
            value = sheet.cell(row=1, column=col).value
            headers.append({
                'column': openpyxl.utils.get_column_letter(col),
                'value': str(value) if value is not None else ""
            })
        return headers

    def _get_samples(self, sheet) -> List[List[str]]:
        """Get sample rows"""
        samples = []
        for row in range(1, min(21, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(26, sheet.max_column + 1)):
                value = sheet.cell(row=row, column=col).value
                row_data.append(str(value)[:200] if value is not None else "")
            samples.append(row_data)
        return samples

    def _collect_sheet_data(self, sheet) -> Dict[str, Any]:
        """Collect comprehensive data from a sheet for analysis"""
        data = {
            'sheet_name': sheet.title,
            'headers': self._get_headers(sheet),
            'samples': self._get_samples(sheet),
            'total_rows': sheet.max_row,
            'total_columns': sheet.max_column
        }

        # Collect all text content for context analysis
        text_content = []
        sections = []

        for row in range(1, min(51, sheet.max_row + 1)):
            row_content = []
            for col in range(1, min(11, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    text_content.append(str(cell_value).strip())
                    row_content.append(str(cell_value).strip())

            if row_content:
                sections.append({
                    'row': row,
                    'content': row_content
                })

        data['text_content'] = text_content[:100]
        data['sections'] = sections[:20]
        return data

    def extract_all(self) -> Dict[str, Any]:
        """Extract questions from all sheets with LLM-based intelligent analysis"""
        print(f"🚀 Enhanced Excel Extraction: {self.file_path}")
        print("=" * 60)

        # Load workbook
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        print(f"📊 Loaded workbook with {len(self.workbook.sheetnames)} sheets")
        print(f"📋 Available sheets: {self.workbook.sheetnames}")

        # STEP 1: LLM-based sheet analysis (determines content vs question sheets)
        sheet_analysis = self._analyze_sheets_with_llm()

        # STEP 2: Extract global context from detected content sheet
        self._extract_global_context_from_detected_sheet()

        # STEP 3: Extract questions from question sheets
        results = {}
        for sheet_name, analysis in sheet_analysis.sheets_analysis.items():
            if not analysis.skip_extraction:
                result = self._extract_from_sheet_enhanced(sheet_name, analysis)
                if result:
                    results[sheet_name] = result

        return {
            'file_path': self.file_path,
            'global_context': self.global_context.model_dump() if self.global_context else None,
            'sheet_results': {name: result.model_dump() for name, result in results.items()},
            'timestamp': datetime.now().isoformat()
        }

    def _analyze_sheets_with_llm(self):
        """LLM-based intelligent sheet analysis"""
        print("\n📊 Step 1: LLM-Based Intelligent Sheet Analysis")
        print("🧠 Analyzing actual content patterns to classify sheets...")

        # Collect data from all sheets
        sheets_info = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            info = {
                'name': sheet_name,
                'rows': sheet.max_row,
                'columns': sheet.max_column,
                'headers': self._get_headers(sheet),
                'sample_data': self._get_samples(sheet)
            }
            sheets_info.append(info)
            print(f"  📋 '{sheet_name}': {sheet.max_row} rows, {sheet.max_column} columns")

        # Use LLM-based analysis
        analysis_result, detected_content_sheet = self.smart_analyzer.analyze_all_sheets_with_llm(sheets_info)

        # Store detected content sheet for later use
        self.detected_content_sheet = detected_content_sheet

        print("\n📊 Final Sheet Classification:")
        for sheet_name, analysis in analysis_result.sheets_analysis.items():
            print(f"  📄 '{sheet_name}': {analysis.sheet_type.value} - Questions: {analysis.contains_questions}")

        return analysis_result

    def _extract_global_context_from_detected_sheet(self):
        """Extract global context from LLM-detected content sheet"""
        print(f"\n📖 Step 2: Global Context Extraction")

        content_sheet_to_use = None

        # Priority 1: Manual override
        if self.content_sheet_name and self.content_sheet_name in self.workbook.sheetnames:
            content_sheet_to_use = self.content_sheet_name
            print(f"📚 Using manually specified content sheet: {self.content_sheet_name}")

        # Priority 2: LLM-detected content sheet
        elif self.detected_content_sheet and self.detected_content_sheet in self.workbook.sheetnames:
            content_sheet_to_use = self.detected_content_sheet
            print(f"🧠 Using LLM-detected content sheet: {self.detected_content_sheet}")

        # Priority 3: First sheet fallback
        elif self.workbook.sheetnames:
            content_sheet_to_use = self.workbook.sheetnames[0]
            print(f"⚠️  No content sheet detected, using first sheet: {content_sheet_to_use}")

        # Extract context from selected sheet
        if content_sheet_to_use:
            sheet = self.workbook[content_sheet_to_use]
            content_data = self._collect_sheet_data(sheet)
            self.global_context = self.claude_client.extract_global_context(content_data)
            print(f"✅ Global context extracted successfully")
            print(f"📚 Document Type: {self.global_context.document_type}")
        else:
            self.global_context = self.claude_client._create_enhanced_fallback_context()
            print("📚 Using fallback global context")

    def _extract_from_sheet_enhanced(self, sheet_name: str, analysis) -> Optional[ExtractionResult]:
        """Enhanced extraction from a single sheet"""
        print(f"\n📋 Step 3: Enhanced Extraction from '{sheet_name}'")

        sheet = self.workbook[sheet_name]

        # Intelligent column detection with statistics
        worksheet_data = {
            'headers': self._get_headers(sheet),
            'samples': self._get_samples(sheet)
        }

        column_info = self.claude_client.detect_columns_with_statistics(worksheet_data, sheet_name)
        print(f"  📊 Question column: {column_info.question_column}")
        print(f"  📋 Answer columns: {column_info.answer_columns} ({len(column_info.answer_columns)} columns)")
        print(f"  🎯 Detection confidence: {column_info.confidence}")

        # Extract questions with enhanced hierarchy parsing
        questions = self._extract_questions_enhanced(sheet, column_info)

        if questions:
            # Calculate comprehensive statistics
            fillable = [q for q in questions if q.should_fill]
            hierarchy_stats = HierarchyStats(
                parent_headers=len([q for q in questions if q.is_parent]),
                numbered_requirements=len([q for q in questions if q.question_type == QuestionType.NUMBERED_REQUIREMENT]),
                lettered_requirements=len([q for q in questions if q.question_type == QuestionType.LETTERED_REQUIREMENT]),
                total_fillable=len(fillable)
            )

            result = ExtractionResult(
                sheet_name=sheet_name,
                document_structure={
                    'total_rows': sheet.max_row,
                    'total_columns': sheet.max_column,
                    'question_columns': [column_info.question_column],
                    'answer_columns': column_info.answer_columns
                },
                total_questions_extracted=len(questions),
                questions=questions,
                statistics={
                    'fillable_questions': len(fillable),
                    'non_fillable_questions': len(questions) - len(fillable),
                    'completion_rate': round((len([q for q in fillable if q.answers]) / max(len(fillable), 1)) * 100, 2)
                },
                hierarchy_stats=hierarchy_stats,
                column_info=column_info
            )

            print(f"  ✅ Extracted {len(questions)} items ({len(fillable)} fillable)")
            print(f"  📊 Hierarchy: {hierarchy_stats.parent_headers} parents, {hierarchy_stats.total_fillable} requirements")
            return result

        return None

    def _extract_questions_enhanced(self, sheet, column_info) -> List[ExtractedQuestion]:
        """Enhanced question extraction with better hierarchy parsing"""
        questions = []
        question_texts = []
        question_id = 1
        multi_line_parser = MultiLineQuestionParser()

        # Extract questions
        question_col_idx = ord(column_info.question_column) - ord('A') + 1

        for row in range(column_info.start_row, sheet.max_row + 1):
            if row in column_info.skip_rows:
                continue

            cell_value = sheet.cell(row=row, column=question_col_idx).value
            if cell_value:
                parsed_content = multi_line_parser.parse_cell_content(str(cell_value).strip())

                if len(parsed_content) > 1:
                    # Multiple sub-questions in one cell
                    parent_text = parsed_content[0]['text']
                    question_texts.append({
                        'row': row,
                        'text': parent_text,
                        'is_sub_question': False
                    })

                    for i, sub_q in enumerate(parsed_content[1:]):
                        question_texts.append({
                            'row': row,
                            'text': sub_q['text'],
                            'is_sub_question': True,
                            'parent_text': parent_text,
                            'sub_index': i + 1
                        })
                else:
                    question_texts.append({
                        'row': row,
                        'text': parsed_content[0]['text'],
                        'is_sub_question': False
                    })

        # Enhanced hierarchy parsing
        main_questions = [q for q in question_texts if not q.get('is_sub_question')]

        if main_questions:
            print(f"    🧠 Enhanced hierarchy parsing for {len(main_questions)} items...")
            try:
                if self.hierarchy_parser:
                    parsed_hierarchy = self.hierarchy_parser.parse_questions_contextually_simplified(
                        main_questions,
                        chunk_size=100,
                        overlap=20
                    )
                else:
                    # Fallback to rule-based parsing
                    parsed_hierarchy = self._rule_based_hierarchy_parsing(main_questions)
            except Exception as e:
                print(f"    ❌ Hierarchy parsing failed: {str(e)}")
                # Fallback
                parsed_hierarchy = [{
                    'question_type': QuestionType.GENERAL_QUESTION,
                    'is_parent': False,
                    'should_fill': True,
                    'hierarchy_level': 0,
                    'parent_id': None,
                    'parent_text': None
                } for _ in main_questions]

            # Create ExtractedQuestion objects
            hierarchy_idx = 0
            for q_data in question_texts:
                if q_data.get('is_sub_question'):
                    hierarchy = {
                        'question_type': QuestionType.SUB_LIST_REQUIREMENT,
                        'is_parent': False,
                        'should_fill': True,
                        'hierarchy_level': 2,
                        'parent_id': None,
                        'parent_text': q_data.get('parent_text')
                    }
                else:
                    if hierarchy_idx < len(parsed_hierarchy):
                        hierarchy = parsed_hierarchy[hierarchy_idx]
                        hierarchy_idx += 1
                    else:
                        hierarchy = {
                            'question_type': QuestionType.GENERAL_QUESTION,
                            'is_parent': False,
                            'should_fill': True,
                            'hierarchy_level': 0,
                            'parent_id': None,
                            'parent_text': None
                        }

                # Extract answers if should_fill
                answers = {}
                if hierarchy['should_fill']:
                    for ans_col in column_info.answer_columns:
                        ans_col_idx = ord(ans_col) - ord('A') + 1
                        cell_value = sheet.cell(row=q_data['row'], column=ans_col_idx).value
                        if cell_value:
                            answers[ans_col] = str(cell_value).strip()

                question = ExtractedQuestion(
                    question_id=question_id,
                    row_id=q_data['row'],
                    column_letter=column_info.question_column,
                    question=q_data['text'],
                    answers=answers,
                    question_type=hierarchy['question_type'],
                    is_parent=hierarchy['is_parent'],
                    should_fill=hierarchy['should_fill'],
                    parent_id=hierarchy.get('parent_id'),
                    parent_text=hierarchy.get('parent_text'),
                    hierarchy_level=hierarchy.get('hierarchy_level')
                )

                questions.append(question)
                question_id += 1

        return questions

    def _rule_based_hierarchy_parsing(self, questions: List[Dict]) -> List[Dict[str, Any]]:
        """Simple rule-based hierarchy parsing as fallback"""
        results = []
        
        for q in questions:
            text = q['text'].strip()
            
            # Check for parent headers (end with : and contain "following")
            if text.endswith(':') and any(word in text.lower() for word in ['following', 'include', 'comprise']):
                result = {
                    'question_type': QuestionType.PARENT_HEADER,
                    'is_parent': True,
                    'should_fill': False,
                    'hierarchy_level': 0,
                    'parent_id': None,
                    'parent_text': None
                }
            # Check for numbered requirements
            elif text.startswith(tuple(f"{i})" for i in range(1, 100))):
                result = {
                    'question_type': QuestionType.NUMBERED_REQUIREMENT,
                    'is_parent': False,
                    'should_fill': True,
                    'hierarchy_level': 1,
                    'parent_id': None,
                    'parent_text': None
                }
            # Check for lettered requirements
            elif text.startswith(tuple(f"{chr(i)}." for i in range(ord('a'), ord('z')+1))):
                result = {
                    'question_type': QuestionType.LETTERED_REQUIREMENT,
                    'is_parent': False,
                    'should_fill': True,
                    'hierarchy_level': 1,
                    'parent_id': None,
                    'parent_text': None
                }
            else:
                result = {
                    'question_type': QuestionType.GENERAL_QUESTION,
                    'is_parent': False,
                    'should_fill': True,
                    'hierarchy_level': 0,
                    'parent_id': None,
                    'parent_text': None
                }
            
            results.append(result)
        
        return results