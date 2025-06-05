"""
Enhanced Excel filler with intelligent cross-column logic
"""

import os
import random
import openpyxl
from typing import Dict, List, Any, Optional
from datetime import datetime

from models import (
    ExtractionResult, ExtractedQuestion, GlobalContext, 
    FillStrategy, ColumnFillStrategy
)
# from claude_client import ClaudeStructuredClient
from gemini_client import GeminiStructuredClient


class EnhancedExcelFiller:
    """Enhanced filler with intelligent cross-column logic"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        # self.claude_client = ClaudeStructuredClient()

        self.gemini_client = GeminiStructuredClient()

    def fill_all(self, extraction_results: Dict[str, Any]) -> str:
        """Enhanced filling with intelligent logic"""
        print(f"🔧 Enhanced Excel Filling: {self.file_path}")
        print("=" * 50)

        # Load workbook
        self.workbook = openpyxl.load_workbook(self.file_path)

        # Get global context
        global_context = None
        if extraction_results.get('global_context'):
            global_context = GlobalContext(**extraction_results['global_context'])
            print(f"📚 Using global context: {global_context.document_type}")

        # Fill each sheet with intelligent logic
        total_filled = 0
        for sheet_name, sheet_result_dict in extraction_results.get('sheet_results', {}).items():
            result = ExtractionResult(**sheet_result_dict)
            filled = self._fill_sheet_intelligently(sheet_name, result, global_context)
            total_filled += filled
            print(f"  ✅ Filled {filled} questions in '{sheet_name}'")

        # Save workbook
        output_file = self._save_enhanced_workbook()
        print(f"\n✅ Enhanced filling completed!")
        print(f"📊 Total filled: {total_filled}")
        print(f"📁 Saved to: {output_file}")

        return output_file

    def _fill_sheet_intelligently(self, sheet_name: str, extraction_result: ExtractionResult,
                                global_context: Optional[GlobalContext]) -> int:
        """Fill sheet with intelligent cross-column logic"""
        sheet = self.workbook[sheet_name]

        # Get fillable questions only
        fillable_questions = [q for q in extraction_result.questions
                             if q.should_fill and not q.answers]

        if not fillable_questions:
            print(f"    ℹ️  No fillable questions in '{sheet_name}'")
            return 0

        print(f"    📝 Filling {len(fillable_questions)} questions with intelligent logic")

        # Generate intelligent strategy
        sheet_info = {
            'sheet_name': sheet_name,
            'fillable_questions': len(fillable_questions),
            'answer_columns': extraction_result.column_info.answer_columns if extraction_result.column_info else [],
            'column_purposes': extraction_result.column_info.column_purposes if extraction_result.column_info else {}
        }

        # strategy = self.claude_client.generate_intelligent_fill_strategy(sheet_info, global_context)

        strategy = self.gemini_client.generate_intelligent_fill_strategy(sheet_info, global_context)

        # Apply intelligent strategy
        filled_count = self._apply_intelligent_strategy(sheet, fillable_questions, strategy)

        return filled_count

    def _apply_intelligent_strategy(self, sheet, questions: List[ExtractedQuestion],
                                  strategy: FillStrategy) -> int:
        """Apply intelligent strategy with cross-column logic"""
        if not questions:
            return 0

        distribution = strategy.distribution
        column_strategies = strategy.column_strategies
        cross_rules = strategy.cross_column_rules

        # Calculate distribution
        total = len(questions)
        positive_count = int(total * distribution.positive / 100)
        negative_count = int(total * distribution.negative / 100)
        partial_count = total - positive_count - negative_count

        # Create assignments
        assignments = (['positive'] * positive_count +
                      ['negative'] * negative_count +
                      ['partial'] * partial_count)
        random.shuffle(assignments)

        filled_count = 0

        for question, response_type in zip(questions, assignments):
            row_id = question.row_id
            row_values = {}

            # Apply intelligent cross-column logic
            for col_letter, col_strategy in column_strategies.items():
                try:
                    col_idx = ord(col_letter) - ord('A') + 1

                    # Skip if column doesn't exist
                    if col_idx > sheet.max_column:
                        continue

                    # Get intelligent value with cross-column awareness
                    value = self._get_intelligent_value_with_logic(
                        col_letter, col_strategy, response_type,
                        row_values, cross_rules, question
                    )

                    if value:
                        # Check if cell is empty before filling
                        current_value = sheet.cell(row=row_id, column=col_idx).value
                        if not current_value:
                            sheet.cell(row=row_id, column=col_idx, value=value)
                            row_values[col_letter] = value

                except Exception as e:
                    print(f"      ⚠️  Error filling {row_id},{col_letter}: {e}")

            if row_values:
                filled_count += 1

        return filled_count

    def _get_intelligent_value_with_logic(self, col_letter: str, col_strategy: ColumnFillStrategy,
                                        response_type: str, row_values: Dict,
                                        cross_rules: List[str], question: ExtractedQuestion) -> str:
        """Get value with intelligent cross-column logic"""

        # Check empty probability
        if random.random() < col_strategy.empty_probability:
            return ''

        # Apply cross-column rules
        for rule in cross_rules:
            if self._should_apply_cross_rule(rule, row_values, col_letter, response_type):
                return self._apply_cross_rule(rule, col_letter, response_type)

        # Apply conditional logic from strategy
        conditional_logic = col_strategy.conditional_logic
        if conditional_logic and self._should_apply_conditional_logic(conditional_logic, response_type, row_values):
            return self._apply_conditional_logic(conditional_logic, col_letter, response_type, col_strategy)

        # Get base values based on response type
        if response_type == 'positive':
            values = col_strategy.positive_values
        elif response_type == 'negative':
            values = col_strategy.negative_values
        else:
            values = col_strategy.partial_values

        return random.choice(values) if values else ''

    def _should_apply_cross_rule(self, rule: str, row_values: Dict, col_letter: str, response_type: str) -> bool:
        """Check if cross-column rule should apply"""
        rule_lower = rule.lower()

        # Example rules
        if 'only one' in rule_lower and 'compliance' in rule_lower:
            # Only one compliance column should be marked
            compliance_cols = ['C', 'D', 'E']  # Common compliance columns
            if col_letter in compliance_cols:
                return True

        if 'if' in rule_lower and 'no' in rule_lower:
            # If response is No, apply special logic
            return response_type == 'negative'

        return False

    def _apply_cross_rule(self, rule: str, col_letter: str, response_type: str) -> str:
        """Apply a cross-column rule"""
        rule_lower = rule.lower()

        if 'only one' in rule_lower and 'compliance' in rule_lower:
            # For compliance matrix - only mark the appropriate column
            compliance_mapping = {
                'positive': {'C': '✓', 'D': '', 'E': ''},
                'negative': {'C': '', 'D': '✓', 'E': ''},
                'partial': {'C': '', 'D': '', 'E': '✓'}
            }

            if col_letter in compliance_mapping.get(response_type, {}):
                return compliance_mapping[response_type].get(col_letter, '')

        if 'not applicable' in rule_lower and response_type == 'negative':
            return 'Not applicable'

        return ''

    def _should_apply_conditional_logic(self, logic: str, response_type: str, row_values: Dict) -> bool:
        """Check if conditional logic should apply"""
        logic_lower = logic.lower()

        if 'if positive' in logic_lower:
            return response_type == 'positive'
        elif 'if negative' in logic_lower:
            return response_type == 'negative'
        elif 'if partial' in logic_lower:
            return response_type == 'partial'

        return False

    def _apply_conditional_logic(self, logic: str, col_letter: str, response_type: str, col_strategy: ColumnFillStrategy) -> str:
        """Apply conditional logic"""
        logic_lower = logic.lower()

        if 'mark' in logic_lower and 'only' in logic_lower:
            # Mark this column only for this response type
            if response_type == 'positive':
                return random.choice(col_strategy.positive_values) if col_strategy.positive_values else ''
            elif response_type == 'negative':
                return random.choice(col_strategy.negative_values) if col_strategy.negative_values else ''
            else:
                return random.choice(col_strategy.partial_values) if col_strategy.partial_values else ''

        return ''

    def _save_enhanced_workbook(self) -> str:
        """Save the enhanced filled workbook"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(self.file_path))[0]
        output_file = f"enhanced_filled_{timestamp}_{base_name}.xlsx"
        self.workbook.save(output_file)
        return output_file